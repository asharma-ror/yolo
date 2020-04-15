import contextlib
from typing import Iterator, List, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet._read_only import ReadOnlyWorksheet  # noqa

from .base import MONTHS


@contextlib.contextmanager
def open_book(file_path):
    doc = load_workbook(filename=file_path, read_only=True, data_only=True)
    try:
        yield doc
    finally:
        doc.close()


Row = Tuple[int, List[object]]


def iter_rows(sheet) -> Iterator[Row]:
    for idx, row in enumerate(sheet.rows):  # noqa
        yield (idx, [cell.value for cell in row])


def get_monthly_sheets(fname) -> Iterator[ReadOnlyWorksheet]:
    with open_book(fname) as book:
        for sheet in book.sheetnames:
            if sheet and sheet.strip().lower() in MONTHS:
                yield sheet, iter_rows(book[sheet])
